import streamlit as st
import tempfile
import os
import shutil
from pathlib import Path
import sys

# 添加当前目录到Python路径，确保能找到模块
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# 尝试导入核心处理器
try:
    from excel_processor_optimized import OptimizedExcelProcessor, ProcessingConfig, ProgressTracker
except ImportError as e:
    st.error(f"❌ 无法导入核心处理器模块: {e}")
    st.error("请确保 excel_processor_optimized.py 文件在同一目录下")
    st.stop()

import pandas as pd
import openpyxl
from copy import copy
import zipfile
import json
import time
import threading
from io import BytesIO

# 页面配置
st.set_page_config(
    page_title="Excel处理工作台", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
    }
    .progress-container {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .file-info {
        background-color: #e8f4fd;
        padding: 0.5rem;
        border-radius: 5px;
        margin: 0.5rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# 标题
st.markdown("""
<div class="main-header">
    <h1>📊 Excel处理自动化工作台</h1>
    <p>🚀 专为大规模数据处理优化 | 支持大文件、多线程、内存管理</p>
</div>
""", unsafe_allow_html=True)

# 功能说明
st.markdown("""
### ✨ 优化特性
- **🚀 高性能处理**：支持大文件（>100MB）、多线程并行处理
- **💾 智能内存管理**：自动内存监控、分块处理、垃圾回收
- **📊 实时进度反馈**：处理进度条、ETA时间估算
- **🔧 可配置参数**：批处理大小、线程数、内存限制
- **📁 完整格式保留**：字体、颜色、边框、对齐、日期格式等100%还原
- **🎯 自定义分组**：支持字段值自定义分组拆分
""")

# 侧边栏配置
with st.sidebar:
    st.header("⚙️ 性能配置")
    
    # 批处理大小
    batch_size = st.slider(
        "批处理大小", 
        min_value=100, 
        max_value=5000, 
        value=1000, 
        step=100,
        help="每次处理的数据行数，大文件建议使用较小值"
    )
    
    # 最大线程数
    max_workers = st.slider(
        "最大线程数", 
        min_value=1, 
        max_value=8, 
        value=4, 
        step=1,
        help="并行处理的线程数，建议不超过CPU核心数"
    )
    
    # 内存限制
    memory_limit_mb = st.slider(
        "内存限制(MB)", 
        min_value=256, 
        max_value=2048, 
        value=512, 
        step=128,
        help="内存使用限制，超过时自动垃圾回收"
    )
    
    # 文件大小警告阈值
    file_size_warning_mb = st.slider(
        "大文件警告阈值(MB)", 
        min_value=10, 
        max_value=100, 
        value=50, 
        step=10,
        help="超过此大小的文件会显示性能提示"
    )

# 进度条容器
progress_container = st.container()

# 操作模式选择
mode = st.radio("请选择操作模式：", ["拆分大表为多个小表", "合并多个小表为大表"])

if mode == "拆分大表为多个小表":
    uploaded_file = st.file_uploader("上传Excel文件", type=["xlsx"], help="支持大文件上传")
    
    if uploaded_file:
        st.write("文件已上传")
        print("文件已上传")
        # 文件信息显示
        file_size = len(uploaded_file.getvalue()) / 1024 / 1024  # MB
        st.markdown(f"""
        <div class="file-info">
            <strong>📁 文件信息：</strong><br>
            文件名：{uploaded_file.name}<br>
            文件大小：{file_size:.2f} MB<br>
            文件类型：Excel (.xlsx)
        </div>
        """, unsafe_allow_html=True)
        
        # 大文件警告
        if file_size > file_size_warning_mb:
            st.markdown(f"""
            <div class="warning-box">
                <strong>⚠️ 大文件检测</strong><br>
                文件大小 {file_size:.2f}MB 超过 {file_size_warning_mb}MB，建议：<br>
                • 使用较小的批处理大小（{batch_size}）<br>
                • 确保有足够的内存空间<br>
                • 处理时间可能较长，请耐心等待
            </div>
            """, unsafe_allow_html=True)
        
        # 保存上传文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        st.write(f"临时文件已保存: {tmp_path}")
        print(f"临时文件已保存: {tmp_path}")
        
        # 读取sheet名时用read_only=True
        sheet_names = None
        try:
            wb_tmp = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb_tmp.sheetnames
            wb_tmp.close()
        except Exception as e:
            st.error(f"读取Excel失败: {e}")
            print(f"读取Excel失败: {e}")
        
        # 只要sheet_names获取成功就渲染sheet选择控件
        if sheet_names:
            selected_sheets = st.multiselect("选择要参与拆分的工作表（可多选）", sheet_names, default=sheet_names)
            keep_fields_dict = {}
            sheet_columns_dict = {}
            for sheet in selected_sheets:
                try:
                    df_sheet = pd.read_excel(tmp_path, sheet_name=sheet, header=0, nrows=1000)
                    all_columns = df_sheet.columns.tolist()
                    sheet_columns_dict[sheet] = all_columns
                    keep_fields_dict[sheet] = st.multiselect(f"{sheet}保留字段（可多选）", all_columns, default=all_columns, key=f"keep_{sheet}")
                except Exception as e:
                    st.error(f"读取sheet {sheet} 字段失败: {e}")
            if selected_sheets:
                try:
                    df = pd.read_excel(tmp_path, sheet_name=selected_sheets[0], header=0, nrows=1000)
                    all_columns = df.columns.tolist()
                    split_field = st.selectbox("选择拆分字段（每个唯一值生成一个Excel文件）", all_columns)
                    sort_fields = st.multiselect("排序字段（可多选）", all_columns)
                except Exception as e:
                    st.error(f"读取字段失败: {e}")
                preserve_format = st.checkbox("保留单元格格式", value=True, help="保留字体、颜色、边框等格式")
                # 统计拆分值
                with st.spinner("正在统计拆分字段的唯一值..."):
                    split_values = set()
                    for sheet in selected_sheets:
                        try:
                            df_sheet = pd.read_excel(tmp_path, sheet_name=sheet, header=0)
                            if split_field in df_sheet.columns:
                                split_values.update(df_sheet[split_field].dropna().unique())
                        except Exception as e:
                            continue
                    split_values = list(split_values)
                # 自定义分组功能等后续逻辑保持不变
                
                # 自定义分组功能
                st.subheader("🎯 自定义分组配置")
                use_custom_groups = st.checkbox(
                    "启用自定义分组", 
                    value=False, 
                    help="将多个字段值合并到同一个Excel文件中"
                )
                
                if use_custom_groups:
                    # 分组配置管理
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("**分组配置管理**")
                        # 导入分组配置
                        uploaded_config = st.file_uploader("导入分组配置", type=["json"], key="config_upload")
                        if uploaded_config:
                            try:
                                config_data = json.load(uploaded_config)
                                st.session_state.groups = config_data.get('groups', {})
                                st.success("分组配置导入成功！")
                            except Exception as e:
                                st.error(f"配置文件格式错误: {str(e)}")
                        
                        # 导出分组配置
                        if 'groups' in st.session_state and st.session_state.groups:
                            config_json = json.dumps({'groups': st.session_state.groups}, ensure_ascii=False, indent=2)
                            st.download_button(
                                "导出分组配置",
                                config_json,
                                file_name="分组配置.json",
                                mime="application/json"
                            )
                    
                    with col2:
                        st.markdown("**快速操作**")
                        if st.button("清空所有分组"):
                            st.session_state.groups = {}
                            st.rerun()
                        
                        if st.button("自动分组（每个值一个组）"):
                            st.session_state.groups = {f"组{i+1}": [str(val)] for i, val in enumerate(split_values)}
                            st.rerun()
                    
                    # 初始化分组
                    if 'groups' not in st.session_state:
                        st.session_state.groups = {}
                    
                    # 分组编辑界面
                    st.markdown("**📝 编辑分组**")
                    
                    # 添加新分组
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        new_group_name = st.text_input("新分组名称", placeholder="如：技术团队、管理团队")
                    with col2:
                        if st.button("添加分组") and new_group_name:
                            if new_group_name not in st.session_state.groups:
                                st.session_state.groups[new_group_name] = []
                                st.rerun()
                            else:
                                st.error("分组名称已存在！")
                    
                    # 显示现有分组
                    if st.session_state.groups:
                        st.markdown("**当前分组：**")
                        for group_name, group_values in st.session_state.groups.items():
                            with st.expander(f"📁 {group_name} ({len(group_values)}个值)"):
                                col1, col2 = st.columns([3, 1])
                                with col1:
                                    if group_values:
                                        st.write("当前值：", ", ".join(map(str, group_values)))
                                    else:
                                        st.write("暂无值")
                                
                                with col2:
                                    if st.button(f"删除分组", key=f"del_{group_name}"):
                                        del st.session_state.groups[group_name]
                                        st.rerun()
                    
                    # 字段值分配
                    st.markdown("**📋 字段值分配**")
                    st.write(f"拆分字段 '{split_field}' 的所有唯一值：")
                    
                    # 显示未分配的值
                    assigned_values = set()
                    for group_values in st.session_state.groups.values():
                        assigned_values.update(group_values)
                    
                    unassigned_values = [val for val in split_values if str(val) not in assigned_values]
                    
                    if unassigned_values:
                        st.warning(f"⚠️ 还有 {len(unassigned_values)} 个值未分配：{', '.join(map(str, unassigned_values))}")
                        
                        # 批量分配界面
                        col1, col2 = st.columns(2)
                        with col1:
                            selected_values = st.multiselect("选择要分配的值", unassigned_values)
                        with col2:
                            if st.session_state.groups:
                                target_group = st.selectbox("选择目标分组", list(st.session_state.groups.keys()))
                                if st.button("添加到分组") and selected_values and target_group:
                                    st.session_state.groups[target_group].extend([str(val) for val in selected_values])
                                    st.rerun()
                            else:
                                st.write("请先创建分组")
                    else:
                        st.success("✅ 所有字段值已分配完毕！")
                
                # 开始处理按钮
                if st.button("🚀 开始拆分", type="primary"):
                    # 创建进度条
                    progress_bar = progress_container.progress(0)
                    status_text = progress_container.empty()
                    
                    def progress_callback(current, total):
                        progress = current / total if total > 0 else 0
                        progress_bar.progress(progress)
                        status_text.text(f"处理进度: {current}/{total} ({progress*100:.1f}%)")
                    
                    try:
                        # 重置计时器
                        from excel_processor_optimized import detailed_timer
                        detailed_timer.timers.clear()
                        detailed_timer.current_timers.clear()
                        
                        with st.spinner("正在初始化处理..."):
                            # 创建配置
                            config = ProcessingConfig(
                                split_field=split_field,
                                keep_fields=keep_fields_dict,
                                sort_fields=sort_fields,
                                output_dir="output",
                                sheet_name=None,
                                selected_sheets=selected_sheets,  # 传递用户选择的sheet列表
                                preserve_format=preserve_format,
                                batch_size=batch_size,
                                max_workers=max_workers,
                                memory_limit_mb=memory_limit_mb
                            )
                            
                            if use_custom_groups and 'groups' in st.session_state and st.session_state.groups:
                                config.custom_groups = st.session_state.groups
                            
                            # 创建处理器
                            processor = OptimizedExcelProcessor(config)
                        
                        # 开始处理
                        start_time = time.time()
                        
                        with st.spinner("正在处理数据..."):
                            result_files = processor.split_excel_optimized(
                                tmp_path, 
                                progress_callback=progress_callback
                            )
                        
                        # 创建ZIP包
                        with st.spinner("正在打包结果..."):
                            if use_custom_groups and st.session_state.groups:
                                zip_name = "自定义分组拆分结果.zip"
                            else:
                                zip_name = f"拆分结果_{split_field}.zip"
                            
                            zip_path = processor.create_zip_archive(result_files, zip_name)
                        
                        # 完成处理
                        end_time = time.time()
                        processing_time = end_time - start_time
                        
                        # 显示结果
                        st.success(f"""
                        ✅ 拆分完成！
                        
                        📊 **处理结果：**
                        - 生成文件数：{len(result_files)} 个
                        - 处理时间：{processing_time:.2f} 秒
                        - 平均速度：{len(result_files)/processing_time:.2f} 文件/秒
                        - 内存使用：{processor.memory_manager.get_memory_usage():.1f} MB
                        """)
                        
                        # 下载按钮
                        with open(zip_path, "rb") as f:
                            st.download_button(
                                f"📥 下载全部拆分结果（{len(result_files)}个文件）", 
                                f, 
                                file_name=zip_name,
                                mime="application/zip"
                            )
                        
                        # 清理缓存
                        processor.cleanup_cache()
                        
                        # 打印详细计时总结
                        from excel_processor_optimized import detailed_timer
                        detailed_timer.print_summary()
                        
                    except Exception as e:
                        st.error(f"处理过程中出现错误: {str(e)}")
                        st.exception(e)
                    
                    finally:
                        # 清理临时文件
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                        progress_bar.empty()
                        status_text.empty()

elif mode == "合并多个小表为大表":
    uploaded_files = st.file_uploader(
        "上传多个Excel文件", 
        type=["xlsx"], 
        accept_multiple_files=True,
        help="可以选择多个Excel文件进行合并"
    )
    
    if uploaded_files:
        # 文件信息显示
        total_size = sum(len(f.getvalue()) for f in uploaded_files) / 1024 / 1024  # MB
        st.markdown(f"""
        <div class="file-info">
            <strong>📁 文件信息：</strong><br>
            文件数量：{len(uploaded_files)} 个<br>
            总大小：{total_size:.2f} MB<br>
            平均大小：{total_size/len(uploaded_files):.2f} MB
        </div>
        """, unsafe_allow_html=True)
        
        # 大文件警告
        if total_size > file_size_warning_mb:
            st.markdown(f"""
            <div class="warning-box">
                <strong>⚠️ 大文件检测</strong><br>
                总文件大小 {total_size:.2f}MB 超过 {file_size_warning_mb}MB，建议：<br>
                • 使用较小的批处理大小（{batch_size}）<br>
                • 确保有足够的内存空间<br>
                • 处理时间可能较长，请耐心等待
            </div>
            """, unsafe_allow_html=True)
        
        # 保存上传文件
        with tempfile.TemporaryDirectory() as tmpdir:
            file_paths = []
            all_columns = set()
            all_sheet_names = set()
            
            with st.spinner("正在分析文件结构..."):
                for up in uploaded_files:
                    file_path = os.path.join(tmpdir, up.name)
                    with open(file_path, "wb") as f:
                        f.write(up.getvalue())
                    
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    all_sheet_names.update(wb.sheetnames)
                    df = pd.read_excel(file_path, sheet_name=wb.sheetnames[0], nrows=100)  # 只读取前100行
                    all_columns.update(df.columns.tolist())
                    file_paths.append(file_path)
                    wb.close()
            
            all_sheet_names = list(all_sheet_names)
            all_columns = list(all_columns)
            
            # 工作表选择
            selected_sheets = st.multiselect(
                "选择要合并的工作表（可多选）", 
                all_sheet_names, 
                default=all_sheet_names,
                help="选择要合并的工作表，可以多选"
            )
            
            if selected_sheets:
                # 多sheet字段选择
                keep_fields_dict = {}
                for sheet in selected_sheets:
                    df_sheet = pd.read_excel(file_paths[0], sheet_name=sheet, nrows=100)
                    all_columns = df_sheet.columns.tolist()
                    keep_fields_dict[sheet] = st.multiselect(
                        f"{sheet} 保留字段（可多选）", 
                        all_columns, 
                        default=all_columns, 
                        key=f"merge_keep_{sheet}",
                        help=f"选择要保留的字段，默认全选"
                    )
                
                # 排序字段选择
                sort_fields = st.multiselect(
                    "排序字段（可多选）", 
                    all_columns,
                    help="选择用于排序的字段，可以多选"
                )
                
                # 格式保留选项
                preserve_format = st.checkbox("保留单元格格式", value=True, help="保留字体、颜色、边框等格式")
                
                # 输出文件名
                output_file = st.text_input("合并后文件名", value="合并结果.xlsx", help="指定合并后的文件名")
                
                # 开始合并按钮
                if st.button("🚀 开始合并", type="primary"):
                    # 创建进度条
                    progress_bar = progress_container.progress(0)
                    status_text = progress_container.empty()
                    
                    def progress_callback(current, total):
                        progress = current / total if total > 0 else 0
                        progress_bar.progress(progress)
                        status_text.text(f"合并进度: {current}/{total} ({progress*100:.1f}%)")
                    
                    try:
                        # 重置计时器
                        from excel_processor_optimized import detailed_timer
                        detailed_timer.timers.clear()
                        detailed_timer.current_timers.clear()
                        
                        with st.spinner("正在初始化合并..."):
                            # 创建配置
                            config = ProcessingConfig(
                                keep_fields=keep_fields_dict,
                                sort_fields=sort_fields,
                                output_dir="output",
                                sheet_name=None,
                                preserve_format=preserve_format,
                                batch_size=batch_size,
                                max_workers=max_workers,
                                memory_limit_mb=memory_limit_mb
                            )
                            
                            # 创建处理器
                            processor = OptimizedExcelProcessor(config)
                        
                        # 开始处理
                        start_time = time.time()
                        
                        with st.spinner("正在合并数据..."):
                            result_file = processor.merge_excel_files_optimized(
                                file_paths, 
                                output_file,
                                progress_callback=progress_callback
                            )
                        
                        # 完成处理
                        end_time = time.time()
                        processing_time = end_time - start_time
                        
                        # 显示结果
                        st.success(f"""
                        ✅ 合并完成！
                        
                        📊 **处理结果：**
                        - 合并文件数：{len(file_paths)} 个
                        - 处理时间：{processing_time:.2f} 秒
                        - 平均速度：{len(file_paths)/processing_time:.2f} 文件/秒
                        - 内存使用：{processor.memory_manager.get_memory_usage():.1f} MB
                        - 输出文件：{output_file}
                        """)
                        
                        # 下载按钮
                        with open(result_file, "rb") as f:
                            st.download_button(
                                f"📥 下载合并结果", 
                                f, 
                                file_name=output_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        # 清理缓存
                        processor.cleanup_cache()
                        
                        # 打印详细计时总结
                        from excel_processor_optimized import detailed_timer
                        detailed_timer.print_summary()
                        
                    except Exception as e:
                        st.error(f"合并过程中出现错误: {str(e)}")
                        st.exception(e)
                    
                    finally:
                        progress_bar.empty()
                        status_text.empty()

# 页脚信息
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666;">
    <p>🚀 Excel处理工作台 - 优化版 | 专为大规模数据处理设计</p>
    <p>支持大文件、多线程、智能内存管理、实时进度反馈</p>
</div>
""", unsafe_allow_html=True) 