#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试自定义分组功能
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from pathlib import Path

def create_test_data():
    """创建测试数据"""
    data = {
        '姓名': ['张三', '李四', '王五', '赵六', '钱七', '孙八', '周九', '吴十'],
        '部门': ['技术部', '研发部', '人事部', '财务部', '销售部', '市场部', 'IT部', '行政部'],
        '职位': ['工程师', '高级工程师', 'HR专员', '会计', '销售经理', '市场专员', '系统管理员', '行政专员'],
        '薪资': [8000, 12000, 6000, 7000, 10000, 8000, 9000, 5000]
    }
    
    df = pd.DataFrame(data)
    
    # 创建带格式的Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工信息"
    
    # 写入数据并设置格式
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置表头格式
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                # 设置数据行格式
                if c_idx == 4:  # 薪资列
                    cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    
    # 保存文件
    wb.save("test_data.xlsx")
    print("测试数据文件已创建: test_data.xlsx")
    return df

def test_custom_groups():
    """测试自定义分组功能"""
    from excel_processor import ExcelProcessor, ProcessingConfig
    
    # 创建测试数据
    create_test_data()
    
    # 定义自定义分组配置
    config = ProcessingConfig(
        split_field="部门",
        keep_fields=["姓名", "部门", "职位", "薪资"],
        sort_fields=["部门", "姓名"],
        output_dir="test_output",
        sheet_name="员工信息",
        preserve_format=True,
        custom_groups={
            "技术团队": ["技术部", "研发部", "IT部"],
            "管理团队": ["人事部", "财务部", "行政部"],
            "销售团队": ["销售部", "市场部"]
        }
    )
    
    # 创建输出目录
    Path("test_output").mkdir(exist_ok=True)
    
    # 执行拆分
    processor = ExcelProcessor(config)
    processor.split_excel("test_data.xlsx")
    
    print("自定义分组拆分测试完成！")
    print("生成的文件：")
    for file in Path("test_output").glob("*.xlsx"):
        print(f"  - {file.name}")

if __name__ == "__main__":
    test_custom_groups() 