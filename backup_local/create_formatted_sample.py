#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建包含各种格式的示例Excel文件
用于测试格式复制功能
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

def create_formatted_sample_data():
    """创建包含各种格式的示例数据"""
    
    # 生成员工数据
    employees = []
    departments = ['销售部', '技术部', '人事部', '财务部', '市场部']
    positions = ['经理', '主管', '专员', '助理']
    
    for i in range(20):
        dept = random.choice(departments)
        position = random.choice(positions)
        hire_date = datetime.now() - timedelta(days=random.randint(100, 1000))
        salary = random.randint(5000, 20000)
        
        employee = {
            '员工编号': f'EMP{str(i+1).zfill(3)}',
            '姓名': f'员工{i+1}',
            '部门': dept,
            '职位': position,
            '入职日期': hire_date,
            '薪资': salary,
            '绩效评分': round(random.uniform(3.0, 5.0), 1),
            '状态': random.choice(['在职', '试用期', '离职']),
            '备注': f'这是员工{i+1}的备注信息'
        }
        employees.append(employee)
    
    return pd.DataFrame(employees)

def apply_comprehensive_formatting(wb, sheet_name):
    """应用全面的格式设置"""
    ws = wb[sheet_name]
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据行样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 特殊样式
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    warning_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
    
    # 应用表头格式
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 应用数据行格式
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 特殊格式设置
            if col == 5:  # 入职日期列
                cell.number_format = 'yyyy-mm-dd'
            elif col == 6:  # 薪资列
                cell.number_format = '#,##0'
                if cell.value and cell.value > 15000:
                    cell.font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
            elif col == 7:  # 绩效评分列
                cell.number_format = '0.0'
                if cell.value and cell.value < 4.0:
                    cell.fill = highlight_fill
            elif col == 8:  # 状态列
                if cell.value == '离职':
                    cell.font = warning_font
    
    # 设置列宽
    column_widths = [12, 8, 12, 10, 12, 10, 12, 10, 20]
    for i, width in enumerate(column_widths):
        if i < ws.max_column:
            ws.column_dimensions[get_column_letter(i + 1)].width = width
    
    # 设置行高
    for row in range(1, ws.max_row + 1):
        if row == 1:
            ws.row_dimensions[row].height = 25  # 表头行高
        else:
            ws.row_dimensions[row].height = 20  # 数据行高
    
    # 添加合并单元格（示例：在最后一行添加一个合并的备注）
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value="总体备注：")
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{last_row}:I{last_row}')
    merged_cell = ws.cell(row=last_row, column=1)
    merged_cell.value = "总体备注：这是一个包含各种格式的示例文件，用于测试格式复制功能"
    merged_cell.alignment = Alignment(horizontal='left', vertical='center')
    merged_cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

def create_formatted_excel():
    """创建包含各种格式的示例Excel文件"""
    
    # 生成数据
    df = create_formatted_sample_data()
    
    # 保存为Excel文件
    output_file = '格式测试文件.xlsx'
    df.to_excel(output_file, index=False, sheet_name='员工信息')
    
    # 应用格式
    wb = openpyxl.load_workbook(output_file)
    apply_comprehensive_formatting(wb, '员工信息')
    wb.save(output_file)
    
    print(f"已创建格式测试文件: {output_file}")
    print(f"包含 {len(df)} 条员工记录")
    print("格式包括：")
    print("- 表头：蓝色背景，白色粗体字，居中对齐")
    print("- 数据行：黑色边框，左对齐")
    print("- 日期列：yyyy-mm-dd格式")
    print("- 薪资列：千分位格式，高薪红色显示")
    print("- 绩效列：一位小数，低分黄色背景")
    print("- 状态列：离职红色显示")
    print("- 合并单元格：最后一行备注")
    print("- 自定义列宽和行高")
    
    return output_file

if __name__ == "__main__":
    create_formatted_excel() 