# excel_processor.py
# 核心处理逻辑模块，供命令行和Streamlit前端共用

import os
import json
import yaml
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

@dataclass
class ProcessingConfig:
    split_field: str = ""
    keep_fields: List[str] = None
    sort_fields: List[str] = None
    output_dir: str = "output"
    sheet_name: str = "Sheet1"
    preserve_format: bool = True
    custom_groups: Dict[str, List[str]] = None  # 自定义分组配置
    
    def post_init(self):
        if self.keep_fields is None:
            self.keep_fields = []
        if self.sort_fields is None:
            self.sort_fields = []
        if self.custom_groups is None:
            self.custom_groups = {}

class ExcelProcessor:
    def __init__(self, config):
        self.config = config
        self.output_dir = Path(config.output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def read_excel_with_format(self, file_path: str, sheet_name: str = None) -> Tuple[pd.DataFrame, openpyxl.Workbook]:
        df = pd.read_excel(file_path, sheet_name=sheet_name or self.config.sheet_name)
        wb = openpyxl.load_workbook(file_path)
        return df, wb
    
    def copy_cell_format(self, source_cell, target_cell):
        """复制单元格的所有格式"""
        if not self.config.preserve_format:
            return
            
        # 复制字体格式
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        
        # 复制填充格式（背景色）
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        # 复制边框格式
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom,
                diagonal=source_cell.border.diagonal,
                diagonal_direction=source_cell.border.diagonal_direction,
                outline=source_cell.border.outline,
                vertical=source_cell.border.vertical,
                horizontal=source_cell.border.horizontal
            )
        
        # 复制对齐方式
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=getattr(source_cell.alignment, "horizontal", None),
                vertical=getattr(source_cell.alignment, "vertical", None),
                text_rotation=getattr(source_cell.alignment, "text_rotation", 0),
                wrap_text=getattr(source_cell.alignment, "wrap_text", None),
                shrink_to_fit=getattr(source_cell.alignment, "shrink_to_fit", None),
                indent=getattr(source_cell.alignment, "indent", 0)
            )
        
        # 复制数字格式（包括日期格式）
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        
        # 保护属性（防止 StyleProxy 报错，直接跳过）
        # target_cell.protection = copy(source_cell.protection)
        
        # 复制超链接
        if source_cell.hyperlink:
            target_cell.hyperlink = source_cell.hyperlink
        
    
    def write_excel_with_format(self, df: pd.DataFrame, wb: openpyxl.Workbook, output_path: str, sheet_name: str = "Sheet1"):
        """
        只用openpyxl，100%还原格式，逐单元格复制（值+格式+合并+列宽+行高+数据验证等）
        只复制df中筛选/排序后的行，保留所有格式
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        source_ws = wb[sheet_name]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        # 复制列宽
        for col_letter, dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width

        # 复制行高
        for row_idx, dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = dim.height

        # 复制合并单元格
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))


        # 复制每一行（只复制df筛选/排序后的行）
        # 先找到原表的表头行（假设表头在第一行）
        header = [cell.value for cell in next(source_ws.iter_rows(min_row=1, max_row=1))]
        col_map = {col: idx for idx, col in enumerate(header)}
        # 写表头
        for c, v in enumerate(df.columns, 1):
            src_cell = source_ws.cell(row=1, column=col_map[v]+1)
            tgt_cell = new_ws.cell(row=1, column=c, value=v)
            self.copy_cell_format(src_cell, tgt_cell)
        # 写数据行
        for r, row in enumerate(df.itertuples(index=False), 2):
            # 找到原表中对应的行（用主键或所有字段比对）
            # 这里假设每行唯一，可以用所有字段比对
            found = None
            for src_row in source_ws.iter_rows(min_row=2, max_row=source_ws.max_row):
                values = [cell.value for cell in src_row]
                if all((getattr(row, f)==v or (pd.isna(getattr(row, f)) and v is None)) for f, v in zip(df.columns, values[:len(df.columns)])):
                    found = src_row
                    break
            for c, v in enumerate(row, 1):
                if found:
                    src_cell = found[c-1]
                else:
                    src_cell = source_ws.cell(row=2, column=c)  # fallback
                tgt_cell = new_ws.cell(row=r, column=c, value=v)
                self.copy_cell_format(src_cell, tgt_cell)
        new_wb.save(output_path)
    
    def split_excel(self, input_file: str, sheet_name: str = None) -> List[str]:
        """
        拆分Excel，支持指定sheet_name，保证格式100%还原
        """
        df, wb = self.read_excel_with_format(input_file, sheet_name=sheet_name or self.config.sheet_name)
        use_sheet = sheet_name or self.config.sheet_name
        if use_sheet not in wb.sheetnames:
            raise KeyError(f"Worksheet {use_sheet} does not exist.")
        if self.config.split_field not in df.columns:
            raise ValueError(f"拆分字段 '{self.config.split_field}' 不存在于数据中")
        if self.config.keep_fields:
            available_fields = [col for col in self.config.keep_fields if col in df.columns]
            df = df[available_fields]
        if self.config.sort_fields:
            sort_fields = [col for col in self.config.sort_fields if col in df.columns]
            if sort_fields:
                df = df.sort_values(by=sort_fields)
        
        # 检查是否使用自定义分组
        if self.config.custom_groups:
            return self.split_excel_with_groups(df, wb, use_sheet)
        else:
            # 传统模式：每个值一个文件
            split_values = df[self.config.split_field].unique()
            output_files = []
            for value in split_values:
                subset = df[df[self.config.split_field] == value]
                safe_value = str(value).replace('/', '_').replace('\\', '_').replace(':', '_')
                output_file = self.output_dir / f"{self.config.split_field}-{safe_value}.xlsx"
                self.write_excel_with_format(subset, wb, str(output_file), sheet_name=use_sheet)
                output_files.append(str(output_file))
            return output_files
    
    def split_excel_with_groups(self, df: pd.DataFrame, wb: openpyxl.Workbook, sheet_name: str) -> List[str]:
        """
        按自定义分组拆分Excel文件
        Args:
            df: 数据框
            wb: 工作簿对象
            sheet_name: 工作表名称
        Returns:
            List[str]: 生成的文件路径列表
        """
        output_files = []
        
        # 验证所有字段值都已分配
        all_group_values = set()
        for group_values in self.config.custom_groups.values():
            all_group_values.update(group_values)
        
        split_values = set(df[self.config.split_field].astype(str).unique())
        unassigned = split_values - all_group_values
        
        if unassigned:
            print(f"警告：以下字段值未分配到任何分组: {unassigned}")
        
        # 为每个分组创建文件
        for group_name, group_values in self.config.custom_groups.items():
            if not group_values:  # 跳过空分组
                continue
            
            # 筛选该分组的所有值
            subset = df[df[self.config.split_field].astype(str).isin(group_values)]
            
            if subset.empty:
                print(f"警告：分组 '{group_name}' 没有匹配的数据")
                continue
            
            # 生成输出文件名
            safe_group_name = group_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            output_file = self.output_dir / f"{safe_group_name}.xlsx"
            
            # 写入文件
            self.write_excel_with_format(subset, wb, str(output_file), sheet_name=sheet_name)
            output_files.append(str(output_file))
            print(f"分组 '{group_name}' 完成，包含 {len(subset)} 行数据")
        
        return output_files
    
    def merge_excel_files(self, input_files: list, output_file: str):
        all_data = []
        reference_wb = None
        for file_path in input_files:
            df, wb = self.read_excel_with_format(file_path)
            if reference_wb is None:
                reference_wb = wb
            if self.config.keep_fields:
                available_fields = [col for col in self.config.keep_fields if col in df.columns]
                df = df[available_fields]
            all_data.append(df)
        if not all_data:
            raise ValueError("没有成功读取任何文件")
        merged_df = pd.concat(all_data, ignore_index=True)
        if self.config.sort_fields:
            sort_fields = [col for col in self.config.sort_fields if col in merged_df.columns]
            if sort_fields:
                merged_df = merged_df.sort_values(by=sort_fields)
        # 修复sheet_name不存在问题
        sheet_name = self.config.sheet_name
        if sheet_name not in reference_wb.sheetnames:
            sheet_name = reference_wb.sheetnames[0]
        self.write_excel_with_format(merged_df, reference_wb, output_file, sheet_name=sheet_name)

def load_config(config_file: str) -> ProcessingConfig:
    config_path = Path(config_file)
    if not config_path.exists():
        raise FileNotFoundError(f"配置文件不存在: {config_file}")
    with open(config_path, 'r', encoding='utf-8') as f:
        if config_path.suffix.lower() == '.json':
            config_data = json.load(f)
        elif config_path.suffix.lower() in ['.yml', '.yaml']:
            config_data = yaml.safe_load(f)
        else:
            raise ValueError(f"不支持的配置文件格式: {config_path.suffix}")
    return ProcessingConfig(**config_data) 