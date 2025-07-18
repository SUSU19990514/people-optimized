#!/usr/bin/env python3
 # -*- coding: utf-8 -*-
"""
Excel处理自动化工作台
HR Excel数据自动拆分与合并的交互式工作台工具

功能特性:
- 按字段值拆分大表为多个小表
- 合并多个小表为一个大表
- 保留所有单元格格式（字体、背景色、边框等）
- 支持配置文件自定义字段筛选、排序
- 支持多sheet处理
- 命令行工具接口
作者: LangGPT
版本: 1.0
"""

import os
import sys
import json
import yaml
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
import argparse
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import typer

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_processor.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class ProcessingConfig:
    """处理配置类"""
    split_field: str = ""  # 拆分字段名
    keep_fields: List[str] = None  # 保留字段列表
    sort_fields: List[str] = None  # 排序字段列表
    output_dir: str = "output"  # 输出目录
    sheet_name: str = "Sheet1"  # 工作表名称
    preserve_format: bool = True  # 是否保留格式
    custom_groups: Dict[str, List[str]] = None  # 自定义分组配置
    
    def post_init(self):
        if self.keep_fields is None:
            self.keep_fields = []
        if self.sort_fields is None:
            self.sort_fields = []
        if self.custom_groups is None:
            self.custom_groups = {}

class ExcelProcessor:
    """Excel处理核心类"""
    
    def init(self, config: ProcessingConfig):
        self.config = config
        self.output_dir = Path(config.output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def read_excel_with_format(self, file_path: str, sheet_name: str = None) -> Tuple[pd.DataFrame, openpyxl.Workbook]:
        """
        读取Excel文件，同时保留格式信息
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，None则读取第一个工作表
        Returns:
            (DataFrame, Workbook): 数据框和工作簿对象
        """
        try:
            # 读取数据
            df = pd.read_excel(file_path, sheet_name=sheet_name or self.config.sheet_name)
            
            # 读取格式
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name or self.config.sheet_name]
            
            logger.info(f"成功读取文件: {file_path}")
            return df, wb
            
        except Exception as e:
            logger.error(f"读取文件失败 {file_path}: {str(e)}")
            raise
    
    def copy_cell_format(self, source_cell, target_cell):
        """复制单元格格式"""
        if not self.config.preserve_format:
            return
            
        # 复制字体
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )
        
        # 复制填充
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        # 复制边框
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        # 复制对齐方式
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
    
    def write_excel_with_format(self, df: pd.DataFrame, wb: openpyxl.Workbook, 
                               output_path: str, sheet_name: str = "Sheet1"):
        """
        写入Excel文件并保留格式
        Args:
            df: 数据框
            wb: 原始工作簿（用于格式参考）
            output_path: 输出路径
            sheet_name: 工作表名称
        """
        try:
            # 创建新的工作簿
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = new_ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 复制格式（如果原工作簿存在对应单元格）
                    if r_idx <= wb[sheet_name].max_row and c_idx <= wb[sheet_name].max_column:
                        source_cell = wb[sheet_name].cell(row=r_idx, column=c_idx)
                        self.copy_cell_format(source_cell, cell)
            
            # 保存文件
            new_wb.save(output_path)
            logger.info(f"成功导出: {output_path}")
            
        except Exception as e:
            logger.error(f"写入文件失败 {output_path}: {str(e)}")
            raise
    
    def split_excel(self, input_file: str):
        """
        按字段值拆分Excel文件
        Args:
            input_file: 输入文件路径
        """
        logger.info(f"开始拆分文件: {input_file}")
        
        # 读取文件
        df, wb = self.read_excel_with_format(input_file)
        
        if self.config.split_field not in df.columns:
            raise ValueError(f"拆分字段 '{self.config.split_field}' 不存在于数据中")
        
        # 字段筛选
        if self.config.keep_fields:
            available_fields = [col for col in self.config.keep_fields if col in df.columns]
            df = df[available_fields]
            logger.info(f"保留字段: {available_fields}")
        
        # 排序
        if self.config.sort_fields:
            sort_fields = [col for col in self.config.sort_fields if col in df.columns]
            if sort_fields:
                df = df.sort_values(by=sort_fields)
                logger.info(f"按字段排序: {sort_fields}")
        
        # 检查是否使用自定义分组
        if self.config.custom_groups:
            self.split_excel_with_groups(df, wb)
        else:
            # 传统模式：按拆分字段分组
            split_values = df[self.config.split_field].unique()
            logger.info(f"发现 {len(split_values)} 个不同的拆分值")
            
            for value in split_values:
                # 筛选数据
                subset = df[df[self.config.split_field] == value]
                
                # 生成输出文件名
                safe_value = str(value).replace('/', '_').replace('\\', '_').replace(':', '_')
                output_file = self.output_dir / f"{self.config.split_field}-{safe_value}.xlsx"
                
                # 写入文件
                self.write_excel_with_format(subset, wb, str(output_file))
            logger.info(f"拆分完成，共生成 {len(split_values)} 个文件")
    
    def split_excel_with_groups(self, df: pd.DataFrame, wb: openpyxl.Workbook):
        """
        按自定义分组拆分Excel文件
        Args:
            df: 数据框
            wb: 工作簿对象
        """
        logger.info(f"开始自定义分组拆分，共 {len(self.config.custom_groups)} 个分组")
        
        # 验证所有字段值都已分配
        all_group_values = set()
        for group_values in self.config.custom_groups.values():
            all_group_values.update(group_values)
        
        split_values = set(df[self.config.split_field].astype(str).unique())
        unassigned = split_values - all_group_values
        
        if unassigned:
            logger.warning(f"以下字段值未分配到任何分组: {unassigned}")
        
        # 为每个分组创建文件
        for group_name, group_values in self.config.custom_groups.items():
            if not group_values:  # 跳过空分组
                continue
            
            # 筛选该分组的所有值
            subset = df[df[self.config.split_field].astype(str).isin(group_values)]
            
            if subset.empty:
                logger.warning(f"分组 '{group_name}' 没有匹配的数据")
                continue
            
            # 生成输出文件名
            safe_group_name = group_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            output_file = self.output_dir / f"{safe_group_name}.xlsx"
            
            # 写入文件
            self.write_excel_with_format(subset, wb, str(output_file))
            logger.info(f"分组 '{group_name}' 完成，包含 {len(subset)} 行数据")
        
        logger.info(f"自定义分组拆分完成，共生成 {len([g for g in self.config.custom_groups.values() if g])} 个文件")
    
    def merge_excel_files(self, input_files: List[str], output_file: str):
        """
        合并多个Excel文件
        Args:
            input_files: 输入文件列表
            output_file: 输出文件路径
        """
        logger.info(f"开始合并 {len(input_files)} 个文件")
        all_data = []
        reference_wb = None
        
        for file_path in input_files:
            try:
                df, wb = self.read_excel_with_format(file_path)
                
                if reference_wb is None:
                    reference_wb = wb
                
                # 字段筛选
                if self.config.keep_fields:
                    available_fields = [col for col in self.config.keep_fields if col in df.columns]
                    df = df[available_fields]
                all_data.append(df)
                logger.info(f"已读取: {file_path} ({len(df)} 行)")
                
            except Exception as e:
                logger.warning(f"跳过文件 {file_path}: {str(e)}")
                continue
        
        if not all_data:
            raise ValueError("没有成功读取任何文件")
        
        # 合并数据
        merged_df = pd.concat(all_data, ignore_index=True)
        
        # 排序
        if self.config.sort_fields:
            sort_fields = [col for col in self.config.sort_fields if col in merged_df.columns]
            if sort_fields:
                merged_df = merged_df.sort_values(by=sort_fields)
                logger.info(f"按字段排序: {sort_fields}")
        
        # 写入合并文件
        if reference_wb:
            self.write_excel_with_format(merged_df, reference_wb, output_file)
        else:
            # 如果没有参考格式，直接保存
            merged_df.to_excel(output_file, index=False)
            logger.info(f"成功导出合并文件: {output_file}")
        logger.info(f"合并完成，共 {len(merged_df)} 行数据")

def load_config(config_file: str) -> ProcessingConfig:
    """从配置文件加载配置"""
    config_path = Path(config_file)
    
    if not config_path.exists():
        raise FileNotFoundError(f"配置文件不存在: {config_file}")
    
    with open(config_path, 'r', encoding='utf-8') as f:
        if config_path.suffix.lower() == '.json':
            config_data = json.load(f)
        elif config_path.suffix.lower() in ['.yml', '.yaml']:
            config_data = yaml.safe_load(f)
        else:
            raise ValueError(f"不支持的配置文件格式: {config_path.suffix}")
    
    return ProcessingConfig(**config_data)

def create_sample_config():
    """创建示例配置文件"""
    config = {
        "split_field": "部门",
        "keep_fields": ["姓名", "部门", "职位", "入职日期", "薪资"],
        "sort_fields": ["部门", "姓名"],
        "output_dir": "output",
        "sheet_name": "Sheet1",
        "preserve_format": True
    }
    
    # 保存JSON配置
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    # 保存YAML配置
    with open('config.yaml', 'w', encoding='utf-8') as f:
        yaml.dump(config, f, default_flow_style=False, allow_unicode=True)
    
    # 创建自定义分组配置示例
    group_config = {
        "split_field": "部门",
        "keep_fields": ["姓名", "部门", "职位", "入职日期", "薪资"],
        "sort_fields": ["部门", "姓名"],
        "output_dir": "output",
        "sheet_name": "Sheet1",
        "preserve_format": True,
        "custom_groups": {
            "技术团队": ["技术部", "研发部", "IT部"],
            "管理团队": ["人事部", "财务部", "行政部"],
            "销售团队": ["销售部", "市场部"]
        }
    }
    
    # 保存分组配置示例
    with open('config_with_groups.json', 'w', encoding='utf-8') as f:
        json.dump(group_config, f, ensure_ascii=False, indent=2)
    
    with open('config_with_groups.yaml', 'w', encoding='utf-8') as f:
        yaml.dump(group_config, f, default_flow_style=False, allow_unicode=True)
    
    logger.info("已创建示例配置文件: config.json, config.yaml, config_with_groups.json, config_with_groups.yaml")

def main():
    """主函数 - 命令行接口"""
    parser = argparse.ArgumentParser(description='Excel处理自动化工作台')
    parser.add_argument('--mode', choices=['split', 'merge'],
                       help='处理模式: split(拆分) 或 merge(合并)')
    parser.add_argument('--config', '-c',
                       help='配置文件路径 (JSON或YAML)。拆分模式支持自定义分组配置')
    parser.add_argument('--input', '-i',
                       help='输入文件路径 (拆分模式) 或文件列表 (合并模式，用逗号分隔)')
    parser.add_argument('--output', '-o',
                       help='输出文件路径 (仅合并模式需要)')
    parser.add_argument('--create-config', action='store_true',
                       help='创建示例配置文件（包括自定义分组配置示例）')
    args = parser.parse_args()
    
    # 先处理创建配置文件逻辑
    if args.create_config:
        create_sample_config()
        return
    
    # 只有没有--create-config时才校验其他参数
    missing = []
    if not args.mode:
        missing.append('--mode')
    if not args.config:
        missing.append('--config/-c')
    if not args.input:
        missing.append('--input/-i')
    if missing:
        parser.error(f'the following arguments are required: {", ".join(missing)}')
    
    try:
        # 加载配置
        config = load_config(args.config)
        
        # 创建处理器
        processor = ExcelProcessor(config)
        
        if args.mode == 'split':
            # 拆分模式
            processor.split_excel(args.input)
            
        elif args.mode == 'merge':
            # 合并模式
            if not args.output:
                raise ValueError("合并模式需要指定输出文件路径 (--output)")
            
            input_files = [f.strip() for f in args.input.split(',')]
            processor.merge_excel_files(input_files, args.output)
    
    except Exception as e:
        logger.error(f"处理失败: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()