#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
示例数据生成器
用于创建测试用的Excel文件
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import random

def create_sample_employee_data():
    """创建示例员工数据"""
    
    # 部门列表
    departments = ['销售部', '技术部', '人事部', '财务部', '市场部']
    
    # 职位列表
    positions = {
        '销售部': ['销售经理', '销售专员', '客户经理', '销售总监'],
        '技术部': ['技术总监', '高级工程师', '工程师', '测试工程师', '产品经理'],
        '人事部': ['人事经理', '招聘专员', '培训专员', '薪酬专员'],
        '财务部': ['财务经理', '会计', '出纳', '财务分析师'],
        '市场部': ['市场经理', '市场专员', '品牌专员', '推广专员']
    }
    
    # 生成员工数据
    employees = []
    for i in range(50):  # 生成50个员工
        dept = random.choice(departments)
        position = random.choice(positions[dept])
        
        # 生成入职日期（过去1-5年）
        days_ago = random.randint(1, 1825)  # 1-5年
        hire_date = datetime.now() - timedelta(days=days_ago)
        
        # 生成薪资（根据职位和部门）
        base_salary = {
            '销售部': 6000,
            '技术部': 12000,
            '人事部': 5000,
            '财务部': 7000,
            '市场部': 8000
        }
        salary = base_salary[dept] + random.randint(-2000, 5000)
        
        employee = {
            '员工编号': f'EMP{str(i+1).zfill(3)}',
            '姓名': f'员工{i+1}',
            '部门': dept,
            '职位': position,
            '入职日期': hire_date.strftime('%Y-%m-%d'),
            '薪资': salary,
            '联系电话': f'138{str(random.randint(10000000, 99999999))}',
            '邮箱': f'employee{i+1}@company.com',
            '状态': random.choice(['在职', '试用期', '离职']),
            '直属上级': f'经理{random.randint(1, 10)}'
        }
        employees.append(employee)
    
    return pd.DataFrame(employees)

def apply_formatting(wb, sheet_name):
    """应用格式到工作表"""
    ws = wb[sheet_name]
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据行样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用表头格式
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 应用数据行格式
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 为薪资列添加特殊格式
            if col == 6:  # 薪资列
                cell.number_format = '#,##0'
    
    # 调整列宽
    column_widths = [10, 8, 12, 15, 12, 10, 15, 20, 10, 12]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width

def create_sample_excel():
    """创建示例Excel文件"""
    
    # 生成数据
    df = create_sample_employee_data()
    
    # 保存为Excel文件
    output_file = '员工花名册.xlsx'
    df.to_excel(output_file, index=False, sheet_name='员工信息')
    
    # 应用格式
    wb = openpyxl.load_workbook(output_file)
    apply_formatting(wb, '员工信息')
    wb.save(output_file)
    
    print(f"已创建示例文件: {output_file}")
    print(f"包含 {len(df)} 条员工记录")
    print(f"部门分布: {df['部门'].value_counts().to_dict()}")
    
    return output_file

if __name__ == "__main__":
    create_sample_excel() 