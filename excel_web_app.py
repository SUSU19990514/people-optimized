import streamlit as st
import tempfile
import os
import shutil
from pathlib import Path
from excel_processor import ExcelProcessor, ProcessingConfig
import pandas as pd
import openpyxl
from copy import copy
import zipfile
import json

st.set_page_config(page_title="Excel处理工作台", layout="wide")
st.title("📊 Excel处理自动化工作台")
st.markdown("""
- 支持Excel拆分与合并，**完整保留所有单元格格式**
- 支持字段筛选、排序、sheet多选、参数可视化配置
- 支持**自定义分组拆分**：将多个字段值合并到同一个Excel文件
- 拆分/合并结果可直接下载
- **格式保留包括**：字体、颜色、边框、对齐、日期格式、数字格式、列宽行高、合并单元格等
""")

mode = st.radio("请选择操作模式：", ["拆分大表为多个小表", "合并多个小表为大表"])

if mode == "拆分大表为多个小表":
    uploaded_file = st.file_uploader("上传", type=["xlsx"])
    if uploaded_file:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        sheet_names = wb.sheetnames
        selected_sheets = st.multiselect("选择要参与拆分的工作表（可多选）", sheet_names, default=sheet_names)
        # 多sheet分别选择保留字段
        keep_fields_dict = {}
        for sheet in selected_sheets:
            df_sheet = pd.read_excel(tmp_path, sheet_name=sheet)
            all_columns = df_sheet.columns.tolist()
            keep_fields_dict[sheet] = st.multiselect(f"{sheet}保留字段（可多选）", all_columns, default=all_columns, key=f"keep_{sheet}")
        # 读取第一个被选中的sheet的字段名做参数配置
        if selected_sheets:
            df = pd.read_excel(tmp_path, sheet_name=selected_sheets[0])
        else:
            df = pd.DataFrame()
        st.dataframe(df.head(10))
        all_columns = df.columns.tolist()
        split_field = st.selectbox("选择拆分字段（每个唯一值生成一个Excel文件）", all_columns)
        sort_fields = st.multiselect("排序字段（可多选）", all_columns)
        preserve_format = st.checkbox("保留单元格格式", value=True)
        
        # 统计所有sheet中split_field的唯一值全集
        split_values = set()
        for sheet in selected_sheets:
            df = pd.read_excel(tmp_path, sheet_name=sheet)
            split_values.update(df[split_field].dropna().unique())
        split_values = list(split_values)
        
        # 自定义分组功能
        st.subheader("🎯 自定义分组配置")
        use_custom_groups = st.checkbox("启用自定义分组", value=False, 
                                       help="将多个字段值合并到同一个Excel文件中")
        
        if use_custom_groups:
            # 分组配置管理
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**分组配置管理**")
                # 导入分组配置
                uploaded_config = st.file_uploader("导入分组配置", type=["json"], key="config_upload")
                if uploaded_config:
                    try:
                        config_data = json.load(uploaded_config)
                        st.session_state.groups = config_data.get('groups', {})
                        st.success("分组配置导入成功！")
                    except Exception as e:
                        st.error(f"配置文件格式错误: {str(e)}")
                
                # 导出分组配置
                if 'groups' in st.session_state and st.session_state.groups:
                    config_json = json.dumps({'groups': st.session_state.groups}, ensure_ascii=False, indent=2)
                    st.download_button(
                        "导出分组配置",
                        config_json,
                        file_name="分组配置.json",
                        mime="application/json"
                    )
            
            with col2:
                st.markdown("**快速操作**")
                if st.button("清空所有分组"):
                    st.session_state.groups = {}
                    st.rerun()
                
                if st.button("自动分组（每个值一个组）"):
                    st.session_state.groups = {f"组{i+1}": [str(val)] for i, val in enumerate(split_values)}
                    st.rerun()
            
            # 初始化分组
            if 'groups' not in st.session_state:
                st.session_state.groups = {}
            
            # 分组编辑界面
            st.markdown("**📝 编辑分组**")
            
            # 添加新分组
            col1, col2 = st.columns([2, 1])
            with col1:
                new_group_name = st.text_input("新分组名称", placeholder="如：技术团队、管理团队")
            with col2:
                if st.button("添加分组") and new_group_name:
                    if new_group_name not in st.session_state.groups:
                        st.session_state.groups[new_group_name] = []
                        st.rerun()
                    else:
                        st.error("分组名称已存在！")
            
            # 显示现有分组
            if st.session_state.groups:
                st.markdown("**当前分组：**")
                for group_name, group_values in st.session_state.groups.items():
                    with st.expander(f"📁 {group_name} ({len(group_values)}个值)"):
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            # 显示当前分组的值
                            if group_values:
                                st.write("当前值：", ", ".join(map(str, group_values)))
                            else:
                                st.write("暂无值")
                        
                        with col2:
                            if st.button(f"删除分组", key=f"del_{group_name}"):
                                del st.session_state.groups[group_name]
                                st.rerun()
            
            # 字段值分配
            st.markdown("**📋 字段值分配**")
            st.write(f"拆分字段 '{split_field}' 的所有唯一值：")
            
            # 显示未分配的值
            assigned_values = set()
            for group_values in st.session_state.groups.values():
                assigned_values.update(group_values)
            
            unassigned_values = [val for val in split_values if str(val) not in assigned_values]
            
            if unassigned_values:
                st.warning(f"⚠️ 还有 {len(unassigned_values)} 个值未分配：{', '.join(map(str, unassigned_values))}")
                
                # 批量分配界面
                col1, col2 = st.columns(2)
                with col1:
                    selected_values = st.multiselect("选择要分配的值", unassigned_values)
                with col2:
                    if st.session_state.groups:
                        target_group = st.selectbox("选择目标分组", list(st.session_state.groups.keys()))
                        if st.button("添加到分组") and selected_values and target_group:
                            st.session_state.groups[target_group].extend([str(val) for val in selected_values])
                            st.rerun()
                    else:
                        st.write("请先创建分组")
            else:
                st.success("✅ 所有字段值已分配完毕！")
        
        if st.button("开始拆分"):
            with st.spinner("正在按拆分字段批量生成Excel..."):
                config = ProcessingConfig(
                    split_field=split_field,
                    keep_fields=keep_fields_dict, # 使用字典传递给ProcessingConfig
                    sort_fields=sort_fields,
                    output_dir="output",
                    sheet_name=None,
                    preserve_format=preserve_format
                )
                wb = openpyxl.load_workbook(tmp_path)
                
                if use_custom_groups and 'groups' in st.session_state and st.session_state.groups:
                    # 自定义分组模式
                    result_files = []
                    for group_name, group_values in st.session_state.groups.items():
                        if not group_values:  # 跳过空分组
                            continue
                        
                        # 创建分组文件
                        safe_group_name = group_name.replace('/', '_').replace('\\', '_').replace(':', '_')
                        out_path = os.path.join("output", f"{safe_group_name}.xlsx")
                        new_wb = openpyxl.Workbook()
                        new_wb.remove(new_wb.active)
                        
                        # 为每个sheet处理该分组的所有值
                        for sheet in selected_sheets:
                            df = pd.read_excel(tmp_path, sheet_name=sheet)
                            # 用各自sheet的保留字段
                            keep_fields = keep_fields_dict.get(sheet, df.columns.tolist())
                            # 筛选该分组的所有值
                            subset = df[df[split_field].astype(str).isin(group_values)]
                            if subset.empty:
                                continue
                            subset = subset[keep_fields]  # 只保留用户选择的字段
                            
                            # 用openpyxl复制格式
                            processor = ExcelProcessor(config)
                            processor.write_excel_with_format(subset, wb, "temp.xlsx", sheet_name=sheet)
                            temp_loaded = openpyxl.load_workbook("temp.xlsx")
                            temp_sheet = temp_loaded.active
                            new_ws = new_wb.create_sheet(title=sheet)
                            
                            # 复制所有格式
                            for row in temp_sheet.iter_rows():
                                for cell in row:
                                    new_cell = new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                                    new_cell.font = copy(cell.font)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.border = copy(cell.border)
                                    new_cell.alignment = copy(cell.alignment)
                                    new_cell.number_format = cell.number_format
                                    new_cell.hyperlink = cell.hyperlink
                            
                            # 复制列宽行高
                            for col_letter, dim in temp_sheet.column_dimensions.items():
                                new_ws.column_dimensions[col_letter].width = dim.width
                            for row_idx, dim in temp_sheet.row_dimensions.items():
                                new_ws.row_dimensions[row_idx].height = dim.height
                            
                            # 复制合并单元格
                            for merged_range in temp_sheet.merged_cells.ranges:
                                new_ws.merge_cells(str(merged_range))
                            
                            # 复制数据验证
                            if hasattr(temp_sheet, 'data_validations') and temp_sheet.data_validations is not None:
                                for dv in temp_sheet.data_validations.dataValidation:
                                    new_ws.add_data_validation(dv)
                            
                            # 复制自动筛选
                            if temp_sheet.auto_filter is not None:
                                new_ws.auto_filter.ref = temp_sheet.auto_filter.ref
                            
                            # 复制页眉页脚
                            new_ws.oddHeader.center.text = temp_sheet.oddHeader.center.text
                            new_ws.oddHeader.left.text = temp_sheet.oddHeader.left.text
                            new_ws.oddHeader.right.text = temp_sheet.oddHeader.right.text
                            new_ws.oddFooter.center.text = temp_sheet.oddFooter.center.text
                            new_ws.oddFooter.left.text = temp_sheet.oddFooter.left.text
                            new_ws.oddFooter.right.text = temp_sheet.oddFooter.right.text
                        
                        new_wb.save(out_path)
                        result_files.append(out_path)
                    
                    # 打包zip
                    zip_path = os.path.join("output", f"自定义分组拆分结果.zip")
                    with zipfile.ZipFile(zip_path, 'w') as zipf:
                        for file in result_files:
                            zipf.write(file, arcname=os.path.basename(file))
                    st.success(f"自定义分组拆分完成，共生成 {len(result_files)} 个Excel文件，已打包为zip")
                    with open(zip_path, "rb") as f:
                        st.download_button(f"下载全部拆分结果（zip）", f, file_name=os.path.basename(zip_path))
                
                else:
                    # 传统模式：每个值一个文件
                    result_files = []
                    for value in split_values:
                        safe_value = str(value).replace('/', '_').replace('\\', '_').replace(':', '_')
                        out_path = os.path.join("output", f"{split_field}-{safe_value}.xlsx")
                        new_wb = openpyxl.Workbook()
                        new_wb.remove(new_wb.active)
                        for sheet in selected_sheets:
                            df = pd.read_excel(tmp_path, sheet_name=sheet)
                            keep_fields = keep_fields_dict.get(sheet, df.columns.tolist())
                            subset = df[df[split_field] == value]
                            if subset.empty:
                                continue
                            subset = subset[keep_fields]  # 只保留用户选择的字段
                            # 用openpyxl复制格式
                            processor = ExcelProcessor(config)
                            processor.write_excel_with_format(subset, wb, "temp.xlsx", sheet_name=sheet)
                            temp_loaded = openpyxl.load_workbook("temp.xlsx")
                            temp_sheet = temp_loaded.active
                            new_ws = new_wb.create_sheet(title=sheet)
                            for row in temp_sheet.iter_rows():
                                for cell in row:
                                    new_cell = new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                                    new_cell.font = copy(cell.font)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.border = copy(cell.border)
                                    new_cell.alignment = copy(cell.alignment)
                                    new_cell.number_format = cell.number_format
                                    new_cell.hyperlink = cell.hyperlink
                            for col_letter, dim in temp_sheet.column_dimensions.items():
                                new_ws.column_dimensions[col_letter].width = dim.width
                            for row_idx, dim in temp_sheet.row_dimensions.items():
                                new_ws.row_dimensions[row_idx].height = dim.height
                            for merged_range in temp_sheet.merged_cells.ranges:
                                new_ws.merge_cells(str(merged_range))
                            if hasattr(temp_sheet, 'data_validations') and temp_sheet.data_validations is not None:
                                for dv in temp_sheet.data_validations.dataValidation:
                                    new_ws.add_data_validation(dv)
                            if temp_sheet.auto_filter is not None:
                                new_ws.auto_filter.ref = temp_sheet.auto_filter.ref
                            new_ws.oddHeader.center.text = temp_sheet.oddHeader.center.text
                            new_ws.oddHeader.left.text = temp_sheet.oddHeader.left.text
                            new_ws.oddHeader.right.text = temp_sheet.oddHeader.right.text
                            new_ws.oddFooter.center.text = temp_sheet.oddFooter.center.text
                            new_ws.oddFooter.left.text = temp_sheet.oddFooter.left.text
                            new_ws.oddFooter.right.text = temp_sheet.oddFooter.right.text
                        new_wb.save(out_path)
                        result_files.append(out_path)
                    # 打包zip
                    zip_path = os.path.join("output", f"拆分结果_{split_field}.zip")
                    with zipfile.ZipFile(zip_path, 'w') as zipf:
                        for file in result_files:
                            zipf.write(file, arcname=os.path.basename(file))
                    st.success(f"拆分完成，共生成 {len(result_files)} 个Excel文件，已打包为zip")
                    with open(zip_path, "rb") as f:
                        st.download_button(f"下载全部拆分结果（zip）", f, file_name=os.path.basename(zip_path))
                
                # 清理临时文件
                os.remove(tmp_path)
                if os.path.exists("temp.xlsx"):
                    os.remove("temp.xlsx")

elif mode == "合并多个小表为大表":
    uploaded_files = st.file_uploader("上传", type=["xlsx"], accept_multiple_files=True)
    if uploaded_files:
        with tempfile.TemporaryDirectory() as tmpdir:
            file_paths = []
            all_columns = set()
            all_sheet_names = set()
            for up in uploaded_files:
                file_path = os.path.join(tmpdir, up.name)
                with open(file_path, "wb") as f:
                    f.write(up.read())
                wb = openpyxl.load_workbook(file_path, read_only=True)
                all_sheet_names.update(wb.sheetnames)
                df = pd.read_excel(file_path, sheet_name=wb.sheetnames[0])
                all_columns.update(df.columns.tolist())
                file_paths.append(file_path)
            all_sheet_names = list(all_sheet_names)
            selected_sheets = st.multiselect("选择要合并的工作表（可多选）", all_sheet_names, default=all_sheet_names)
            all_columns = list(all_columns)
            keep_fields_dict = {}
            for sheet in selected_sheets:
                df_sheet = pd.read_excel(file_paths[0], sheet_name=sheet) # 假设第一个文件的sheet结构代表所有文件
                all_columns = df_sheet.columns.tolist()
                keep_fields_dict[sheet] = st.multiselect(f"{sheet}保留字段（可多选）", all_columns, default=all_columns, key=f"merge_keep_{sheet}")
            sort_fields = st.multiselect("排序字段（可多选）", all_columns)
            preserve_format = st.checkbox("保留单元格格式", value=True)
            output_file = st.text_input("合并后文件名", value="合并结果.xlsx")
            if st.button("开始合并"):
                with st.spinner("正在合并多个sheet..."):
                    config = ProcessingConfig(
                        keep_fields=keep_fields_dict, # 使用字典传递给ProcessingConfig
                        sort_fields=sort_fields,
                        output_dir="output",
                        sheet_name=None,
                        preserve_format=preserve_format
                    )
                    new_wb = openpyxl.Workbook()
                    new_wb.remove(new_wb.active)
                    for sheet in selected_sheets:
                        # 合并所有文件的该sheet
                        dfs = []
                        ref_wb = None
                        for file_path in file_paths:
                            wb = openpyxl.load_workbook(file_path)
                            if sheet in wb.sheetnames:
                                df = pd.read_excel(file_path, sheet_name=sheet)
                                # 用各自sheet的保留字段
                                keep_fields = keep_fields_dict.get(sheet, df.columns.tolist())
                                if keep_fields:
                                    available_fields = [col for col in keep_fields if col in df.columns]
                                    df = df[available_fields]
                                if sort_fields:
                                    sort_fields_valid = [col for col in sort_fields if col in df.columns]
                                    if sort_fields_valid:
                                        df = df.sort_values(by=sort_fields_valid)
                                dfs.append(df)
                                if ref_wb is None:
                                    ref_wb = wb
                        if not dfs:
                            continue
                        merged_df = pd.concat(dfs, ignore_index=True)
                        # 用openpyxl复制格式
                        processor = ExcelProcessor(config)
                        processor.write_excel_with_format(merged_df, ref_wb, "temp_merge.xlsx", sheet_name=sheet)
                        temp_loaded = openpyxl.load_workbook("temp_merge.xlsx")
                        temp_sheet = temp_loaded.active
                        new_ws = new_wb.create_sheet(title=sheet)
                        for row in temp_sheet.iter_rows():
                            for cell in row:
                                new_cell = new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                                new_cell.font = copy(cell.font)
                                new_cell.fill = copy(cell.fill)
                                new_cell.border = copy(cell.border)
                                new_cell.alignment = copy(cell.alignment)
                                new_cell.number_format = cell.number_format
                                new_cell.hyperlink = cell.hyperlink
                        for col_letter, dim in temp_sheet.column_dimensions.items():
                            new_ws.column_dimensions[col_letter].width = dim.width
                        for row_idx, dim in temp_sheet.row_dimensions.items():
                            new_ws.row_dimensions[row_idx].height = dim.height
                        for merged_range in temp_sheet.merged_cells.ranges:
                            new_ws.merge_cells(str(merged_range))
                        if hasattr(temp_sheet, 'data_validations') and temp_sheet.data_validations is not None:
                            for dv in temp_sheet.data_validations.dataValidation:
                                new_ws.add_data_validation(dv)
                        if temp_sheet.auto_filter is not None:
                            new_ws.auto_filter.ref = temp_sheet.auto_filter.ref
                        new_ws.oddHeader.center.text = temp_sheet.oddHeader.center.text
                        new_ws.oddHeader.left.text = temp_sheet.oddHeader.left.text
                        new_ws.oddHeader.right.text = temp_sheet.oddHeader.right.text
                        new_ws.oddFooter.center.text = temp_sheet.oddFooter.center.text
                        new_ws.oddFooter.left.text = temp_sheet.oddFooter.left.text
                        new_ws.oddFooter.right.text = temp_sheet.oddFooter.right.text
                    out_path = os.path.join("output", output_file)
                    new_wb.save(out_path)
                    st.success(f"合并完成，所有sheet已合并到 {output_file}")
                    with open(out_path, "rb") as f:
                        st.download_button(f"下载合并结果", f, file_name=output_file)
                    if os.path.exists("temp_merge.xlsx"):
                        os.remove("temp_merge.xlsx") 